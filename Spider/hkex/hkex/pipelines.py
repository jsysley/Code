# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html
from openpyxl import Workbook
from openpyxl import load_workbook


class HkexPipeline(object):
    def __init__(self):
        self.wb = Workbook()

    def open_spider(self, spider):
        print('open file')
        print self.wb.get_sheet_names()

    def process_item(self, item, spider):
        # print item
        # wb = load_workbook(item.code)
        if item['NO'] in self.wb.get_sheet_names():
            # print('item.NO has created')
            ws = self.wb.get_sheet_by_name(item['NO'])
        else:
            # print('item.NO not created')
            ws = self.wb.create_sheet(item['NO'])
            ws.append(['日期', '持股数', '占比'])
        ws.append([item['date'], item['unit'], item['percent']])
        return item

    def close_spider(self, spider):
        print('save file')
        self.wb.save('02188.xlsx')
